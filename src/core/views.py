# core/views.py

# Imports de Django
from django.forms import ValidationError 
from django.utils import timezone 
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q # Para el filtro en exportar_lista_empleados_excel
from django.utils.encoding import escape_uri_path # Para manejar nombres de archivo
import traceback # Para un mejor manejo de errores en debug

#Import pdf y excel de admin
from django.conf import settings
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
# Imports de librerías externas
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, colors
from openpyxl.utils import get_column_letter

# Imports de tus propios archivos de la aplicación
from .services import (
    autenticar_usuario, crear_empleado_service, crear_horario_service, 
    obtener_roles_service, actualizar_horarios_empleado_service,
    asignar_rol_service,
    eliminar_admin_service,
    obtener_admin_por_id_service,
    actualizar_datos_basicos_empleado_service # Si se usa en otra vista
)
from .models import Sucursal, Horario, Empleado, AsignacionHorario, DiaSemana
from .main import generar_reporte_completo, generar_reporte_detalle_completo, generar_datos_dashboard_general,generar_datos_dashboard_31pte,generar_datos_dashboard_villas,generar_datos_dashboard_nave
# Asegurar la importación de Q si no se hace al inicio

# =======================================================
# === VISTAS PRINCIPALES ===
# =======================================================
def inicio(request):
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login') 

def login_view(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = autenticar_usuario(request, email, password)

        if user is not None:
            login(request, user)

             # Redirección según grupo
            if user.groups.filter(name="Admin").exists():
                return redirect("admin_page")
            elif user.groups.filter(name="Manager").exists():
                return redirect("manager-page")
            else:
                messages.error(request, "No tienes permisos para ingresar.")
        else:
            messages.error(request, "Correo o contraseña incorrectos.")
        return redirect("login")
    
    return render(request, "login.html")

@login_required
def admin_page(request, empleado_id=None):
    if request.method == "POST":
        resultado = asignar_rol_service(request.POST)
        if "error" in resultado:
            messages.error(request, resultado["error"])
        else:
            messages.success(request, resultado["success"])
        return redirect("admin_page") 
    
    administradores = obtener_roles_service()

    admin_editar = None
    if empleado_id:
        admin_editar = obtener_admin_por_id_service(empleado_id)

    return render(
        request, "admin_inicio.html",
        {"administradores": administradores, "admin_editar": admin_editar},
    )

@login_required
def eliminar_admin(request, empleado_id):
    resultado = eliminar_admin_service(empleado_id)

    if "error" in resultado:
        messages.error(request, resultado["error"])
    else:
        messages.success(request, resultado["success"])

    return redirect("admin_page")

@login_required
def manager_page(request):
    return render(request, "manager_inicio.html")


# =======================================================
# 7. VISTA PRINCIPAL DE GESTIÓN (MUESTRA ACTIVOS Y ELIMINADOS)
# =======================================================
@login_required
def gestion_empleados(request):
    # CRUCIAL: Usar Empleado.all_objects para listar activos y eliminados.
    empleados = Empleado.all_objects.all() 
    sucursales = Sucursal.objects.all()
    horarios = Horario.objects.all()
    return render(request, "gestion_empleados.html", {
        "empleados": empleados,
        "sucursales": sucursales,
        "horarios": horarios,
    })

# =======================================================
# 8. VISTA DE EXPORTACIÓN A EXCEL (SOLO EMPLEADOS ACTIVOS)
# =======================================================
@login_required
def exportar_lista_empleados_excel(request):
    try:
        search_term = request.GET.get('q', '').strip()
        
        empleados = Empleado.all_objects.all().order_by('empleado_id')

        if search_term:
            empleados = empleados.filter(
                Q(nombre__icontains=search_term) |
                Q(apellido_paterno__icontains=search_term) |
                Q(apellido_materno__icontains=search_term) |
                Q(email__icontains=search_term) |
                Q(codigo_frappe__icontains=search_term) |
                Q(codigo_checador__icontains=search_term)
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte RH"

        # =========================================
        # 🎨 DEFINICIÓN DE ESTILOS (Fondo Blanco)
        # =========================================
        standard_font = Font(name='Calibri', size=11)
        
        # 1. ENCABEZADO (Azul Claro)
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name='Calibri', size=11)
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # (Eliminamos row_even_fill para que sea todo blanco)

        # 2. SEMÁFORO DE ESTATUS
        status_active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        status_active_font = Font(color="006100", name='Calibri')
        
        status_suspend_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        status_suspend_font = Font(color="9C6500", name='Calibri')
        
        status_baja_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        status_baja_font = Font(color="9C0006", name='Calibri')

        # =========================================
        # ENCABEZADOS
        # =========================================
        headers = [
            'ID', 'Cód. Frappe', 'Cód. Checador', 'Nombre', 'Apellido Paterno', 
            'Apellido Materno', 'Email', 'Sucursal(es)', 'Resumen de Horarios', 
            'Usuario Sistema', 'Estatus', 'Fecha Baja', 'Baja Por'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # =========================================
        # PROCESAR DATOS
        # =========================================
        for i, emp in enumerate(empleados, start=2):
            
            # --- A) Datos ---
            if emp.is_deleted: estatus = "BAJA DEFINITIVA"
            elif emp.user and not emp.user.is_active: estatus = "SUSPENDIDO (Bloqueo)"
            else: estatus = "ACTIVO"

            fecha_baja = emp.deleted_at.strftime('%d/%m/%Y') if emp.deleted_at else '-'
            baja_por = emp.deleted_by if emp.deleted_by else '-'
            usuario_sis = emp.user.username if emp.user else ''

            asignaciones = emp.asignaciones.select_related('sucursal', 'horario', 'dia_especifico', 'tipo_turno').all()
            sucursales_set = set()
            detalles_horario = []

            if not asignaciones:
                detalles_horario.append("Sin horario asignado")
            else:
                for a in asignaciones:
                    if not a.sucursal: continue
                    sucursales_set.add(a.sucursal.nombre_sucursal)
                    
                    horas_txt = "??-??"
                    if a.horario:
                        horas_txt = f"{a.horario.hora_entrada.strftime('%H:%M')}-{a.horario.hora_salida.strftime('%H:%M')}"
                    elif a.hora_entrada_especifica and a.hora_salida_especifica:
                        horas_txt = f"{a.hora_entrada_especifica.strftime('%H:%M')}-{a.hora_salida_especifica.strftime('%H:%M')}"
                    
                    dias_txt = ""
                    if a.dia_especifico: dias_txt = f"({a.dia_especifico.nombre_dia})"
                    elif a.tipo_turno: dias_txt = f"[{a.tipo_turno.descripcion}]"
                    else: dias_txt = "[General]"

                    detalles_horario.append(f"• {a.sucursal.nombre_sucursal[:15]}: {horas_txt} {dias_txt}")

            sucursales_str = ", ".join(sucursales_set) if sucursales_set else "-"
            celda_horario_final = "\n".join(detalles_horario)

            # --- B) Fila ---
            row_data = [
                emp.empleado_id, emp.codigo_frappe, emp.codigo_checador, emp.nombre,
                emp.apellido_paterno, emp.apellido_materno or '', emp.email or '',
                sucursales_str, celda_horario_final, usuario_sis,
                estatus, fecha_baja, baja_por
            ]
            ws.append(row_data)
            current_row = ws.max_row

            # --- C) Estilos (Sin fondo gris) ---
            for col_idx, cell in enumerate(ws[current_row], start=1):
                cell.font = standard_font
                cell.border = thin_border
                # Ya NO aplicamos cell.fill = row_even_fill

                if col_idx in [4, 5, 6, 7]: 
                     cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col_idx == 9: 
                     cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                else: 
                     cell.alignment = Alignment(horizontal='center', vertical='center')

            # Semáforo Estatus
            status_cell = ws.cell(row=current_row, column=11)
            if status_cell.value == "ACTIVO":
                status_cell.fill = status_active_fill
                status_cell.font = status_active_font
            elif "SUSPENDIDO" in status_cell.value:
                status_cell.fill = status_suspend_fill
                status_cell.font = status_suspend_font
            elif "BAJA" in status_cell.value:
                status_cell.fill = status_baja_fill
                status_cell.font = status_baja_font
    
        # Ajustar anchos
        column_widths = [6, 9, 10, 18, 15, 15, 25, 18, 55, 15, 15, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A2'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path("Reporte_RH_Profesional.xlsx")}"'
        return response

    except Exception as e:
        traceback.print_exc()
        messages.error(request, f"Error al generar Excel: {str(e)}")
        return redirect('admin-gestion-empleados')


# =======================================================
# 9. VISTA DE CREACIÓN (Mantenemos tu lógica original)
# =======================================================
@login_required
@require_http_methods(["GET", "POST"])
def crear_empleado(request):
    if request.method == "POST":
        try:
            empleado = crear_empleado_service(request.POST)
            messages.success(request, f"Empleado {empleado.nombre} creado correctamente.")
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, "Ocurrió un error inesperado al crear el empleado.")

        return redirect('admin-gestion-empleados')
    return render(request, "gestion_empleados.html")

# =======================================================
# 10. VISTA DE ELIMINACIÓN (CORREGIDA: Usa el método .delete() del modelo)
# =======================================================
@login_required
def eliminar_empleado(request, empleado_id):
    # CRUCIAL: Usamos all_objects para encontrar el registro sin importar su estado
    empleado = get_object_or_404(Empleado.all_objects, empleado_id=empleado_id)

    # 💥 USO DEL MÉTODO SOBREESCRITO
    # Pasamos el usuario para registrar quién eliminó el registro.
    empleado.delete(user=request.user) 

    messages.success(request, f"El empleado {empleado.nombre} ha sido eliminado lógicamente.")
    return redirect("admin-gestion-empleados")


# =======================================================
# 11. NUEVA VISTA DE RESTAURACIÓN (Para el botón 'Revertir')
# =======================================================
@login_required
def restaurar_empleado(request, empleado_id):
    try:
        # Usamos all_objects para encontrar el registro, incluso si está eliminado
        empleado = get_object_or_404(Empleado.all_objects, empleado_id=empleado_id)
        
        # 💥 USO DEL MÉTODO UNDELETE CREADO EN EL MODELO
        empleado.undelete()
        
        messages.success(request, f"El empleado {empleado.nombre} ha sido restaurado correctamente.")
    except Empleado.DoesNotExist:
        messages.error(request, "Error: El empleado no existe.")
        
    return redirect("admin-gestion-empleados")

# =======================================================
# 12. RESTO DE VISTAS DE EDICIÓN Y API (Mantenemos tu código original)
# =======================================================
@login_required
@require_http_methods(["POST"])
def editar_empleado(request, empleado_id):
    try:
        empleado = actualizar_horarios_empleado_service(empleado_id, request.POST)
        messages.success(request, f"Horarios de {empleado.nombre} actualizados correctamente.")
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, "Ocurrió un error inesperado al actualizar los horarios.")
    return redirect('admin-gestion-empleados')

@login_required
@require_http_methods(["POST"])
def editar_datos_basicos_empleado(request, empleado_id):
    try:
        empleado = actualizar_datos_basicos_empleado_service(empleado_id, request.POST)
        messages.success(request, f"Datos de {empleado.nombre} actualizados correctamente.")
    except ValidationError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error al actualizar datos: {e}")
    return redirect('admin-gestion-empleados')

@login_required
@require_http_methods(["POST"])
def crear_horario(request):
    if request.method == "POST":
        try:
            crear_horario_service(request.POST)
            return JsonResponse({"success": "Horario creado correctamente"}, status=200) 
        except ValidationError as e: 
            return JsonResponse({"error": str(e)}, status=400)
        except Exception as e: 
            return JsonResponse({"error": f"Error interno: {str(e)}"}, status=500)
    return redirect("admin-gestion-empleados")

@login_required
def gestion_usuarios(request):
    return render(request, "gestion_usuarios.html")

@login_required
def reporte_horas(request):
    sucursales = Sucursal.objects.all()
    return render(request, "reporte_horas.html", {"sucursales": sucursales})

@login_required
def lista_asistencias(request):
    sucursales = Sucursal.objects.all()
    return render(request, "lista_asistencias.html", {"sucursales": sucursales})

@require_http_methods(["GET"])
def health_check(request):
    return JsonResponse({'status': 'healthy'}, status=200)

@login_required
@require_http_methods(["GET"])
def api_reporte_horas(request):
    try:
        start_date = request.GET.get("startDate")
        end_date = request.GET.get("endDate") 
        sucursal = request.GET.get("sucursal", "Todas")
        if not start_date or not end_date:
            return JsonResponse({"error": "Debe proporcionar fecha de inicio y fin."}, status=400)
        resultado = generar_reporte_completo(start_date=start_date, end_date=end_date, sucursal=sucursal)
        return JsonResponse(resultado)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error interno del servidor: {str(e)}"}, status=500)

@login_required
@require_http_methods(["GET"])
def api_reporte_detalle(request):
    try:
        start_date = request.GET.get("startDate")
        end_date = request.GET.get("endDate")
        sucursal = request.GET.get("sucursal", "Todas")
        if not start_date or not end_date:
            return JsonResponse({"success": False, "error": "Debe proporcionar fecha de inicio y fin."}, status=400)
        resultado = generar_reporte_detalle_completo(start_date=start_date, end_date=end_date, sucursal=sucursal)
        return JsonResponse(resultado)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error interno del servidor: {str(e)}"}, status=500)

@login_required
def grafica_general(request):
    is_admin = request.user.groups.filter(name="Admin").exists()

    context = {
        'is_admin': is_admin, 
    }
    return render(request, "grafica_general.html", context) 

@login_required
@require_http_methods(["GET"])
def api_dashboard_general(request):
    try:
        start_date = request.GET.get("startDate")
        end_date = request.GET.get("endDate")
        if not start_date or not end_date:
            return JsonResponse({"success": False, "error": "Fechas de inicio y fin son requeridas."}, status=400)
        
        resultado = generar_datos_dashboard_general(start_date=start_date, end_date=end_date)
        return JsonResponse(resultado)

    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error interno del servidor: {str(e)}"}, status=500)

@login_required
@require_http_methods(["POST"]) 
def exportar_excel_con_colores(request):
    try:
        data = json.loads(request.body)
        nombre_archivo = data.get('nombre_archivo', 'reporte')
        sheets_data = data.get('sheets', {})

        if not sheets_data:
            return JsonResponse({'error': 'No hay datos de hojas para exportar'}, status=400)

        wb = openpyxl.Workbook()
        wb.remove(wb.active) 

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="8DB4E3", end_color="8DB4E3", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal='center', vertical='center')
        
        color_map = {
            'fila-ok': '92D050', 'fila-retardo-normal': 'FFFF00', 'fila-falta': 'FF0000', 
            'fila-descanso': 'B2A1C7', 'fila-permiso': 'FFC0CB', 'fila-retardo-mayor': 'FF9900',
            'fila-salida-anticipada': 'E76F51', 'fila-retardo-cumplido': 'E6D4ED', 
            'fila-totales': 'DDEBF7'
        }

        for sheet_name, sheet_content in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name.capitalize())
            
            datos = sheet_content.get('datos', [])
            colores_clases = sheet_content.get('colores', []) 

            if not datos: continue

            ws.append(datos[0])
            for cell in ws[1]:
                cell.border = thin_border
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            for i, row_data in enumerate(datos[1:]): 
                ws.append(row_data)
                clase_color = colores_clases[i] if i < len(colores_clases) else None
                fill_color = None
                if clase_color in color_map:
                    fill_color = PatternFill(start_color=color_map[clase_color], end_color=color_map[clase_color], fill_type="solid")

                for cell in ws[ws.max_row]: 
                    cell.border = thin_border
                    if fill_color: cell.fill = fill_color
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 4), 50) 
                ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
        return response
        
    except Exception as e:
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def export_dashboard_excel(request):
    if request.method == 'POST':
        try:
            data_from_js = json.loads(request.body)
            branches_data = data_from_js.get("branches", [])
            employee_summary_data = data_from_js.get("employee_summary_kpis", [])
            employee_kpis_data = data_from_js.get("employee_performance_kpis", [])
            
            if not branches_data and not employee_summary_data and not employee_kpis_data:
                return JsonResponse({"error": "No hay datos para exportar."}, status=404)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            colors = {
                'headerBg': '2F5496', 'headerFont': 'FFFFFF', 'successFill': 'C6EFCE', 'successFont': '006100',
                'warningFill': 'FFEB9C', 'warningFont': '9C6500', 'dangerFill': 'FFC7CE', 'dangerFont': '9C0006',
                'totalFill': 'F0F0F0', 'legendHeaderFill': 'BFBFBF', 'legendFont': '595959'
            }
            
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            
            header_font = Font(color=colors['headerFont'], bold=True, size=11)
            success_font = Font(color=colors['successFont']); warning_font = Font(color=colors['warningFont'])
            danger_font = Font(color=colors['dangerFont']); total_font = Font(bold=True)
            legend_header_font = Font(bold=True, size=12); legend_text_font = Font(italic=True, size=10, color=colors['legendFont'])
            
            header_fill = PatternFill(start_color=colors['headerBg'], end_color=colors['headerBg'], fill_type="solid")
            success_fill = PatternFill(start_color=colors['successFill'], end_color=colors['successFill'], fill_type="solid")
            warning_fill = PatternFill(start_color=colors['warningFill'], end_color=colors['warningFill'], fill_type="solid")
            danger_fill = PatternFill(start_color=colors['dangerFill'], end_color=colors['dangerFill'], fill_type="solid")
            total_fill = PatternFill(start_color=colors['totalFill'], end_color=colors['totalFill'], fill_type="solid")
            legend_header_fill = PatternFill(start_color=colors['legendHeaderFill'], end_color=colors['legendHeaderFill'], fill_type="solid")

            number_format_percent_2dec = '0.00%' ; number_format_percent_1dec = '0.0%'
            number_format_decimal_2dec = '0.00'; number_format_integer = '0';
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


            if branches_data:
                ws1 = wb.create_sheet(title="Resumen Sucursal")
                headers1 = ['Sucursal', 'Empleados', 'Eficiencia (%)', 'Puntualidad (%)', 'SIC Promedio', 'Ausencias']
                ws1.append(headers1)
                
                for col_idx, header in enumerate(headers1, 1):
                    cell = ws1.cell(row=1, column=col_idx)
                    cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align; cell.border = thin_border
                    
                num_branches = len(branches_data)
                total_employees = sum(b.get('employees', 0) for b in branches_data)
                total_absences = sum(b.get('absences', 0) for b in branches_data)
                avg_efficiency = sum(b.get('efficiency', 0) for b in branches_data) / num_branches if num_branches > 0 else 0
                avg_punctuality = sum(b.get('punctuality', 0) for b in branches_data) / num_branches if num_branches > 0 else 0
                avg_sic = sum(b.get('avgSIC', 0) for b in branches_data) / num_branches if num_branches > 0 else 0
                
                for row_idx, branch in enumerate(branches_data, 2):
                    ws1.cell(row=row_idx, column=1, value=branch.get('name', 'N/A'))
                    ws1.cell(row=row_idx, column=2, value=branch.get('employees', 0)).number_format = number_format_integer
                    
                    c3 = ws1.cell(row=row_idx, column=3, value=branch.get('efficiency', 0) / 100)
                    c3.number_format = number_format_percent_1dec
                    
                    c4 = ws1.cell(row=row_idx, column=4, value=branch.get('punctuality', 0) / 100)
                    c4.number_format = number_format_percent_1dec
                    
                    c5 = ws1.cell(row=row_idx, column=5, value=branch.get('avgSIC', 0))
                    c5.number_format = number_format_decimal_2dec
                    
                    ws1.cell(row=row_idx, column=6, value=branch.get('absences', 0)).number_format = number_format_integer
                    
                    eff_val = c3.value; pun_val = c4.value; sic_val = c5.value
                    if eff_val is not None:
                        if eff_val >= 0.95: c3.fill = success_fill; c3.font = success_font
                        elif eff_val >= 0.85: c3.fill = warning_fill; c3.font = warning_font
                        else: c3.fill = danger_fill; c3.font = danger_font
                    
                    if pun_val is not None:
                        if pun_val >= 0.95: c4.fill = success_fill; c4.font = success_font
                        elif pun_val >= 0.85: c4.fill = warning_fill; c4.font = warning_font
                        else: c4.fill = danger_fill; c4.font = danger_font
                    
                    if sic_val is not None:
                        if sic_val >= 85: c5.fill = success_fill; c5.font = success_font
                        elif sic_val >= 50: c5.fill = warning_fill; c5.font = warning_font
                        else: c5.fill = danger_fill; c5.font = danger_font
                        
                    for c_idx in range(1, len(headers1) + 1):
                        ws1.cell(row=row_idx, column=c_idx).border = thin_border
                        
                total_row_idx = len(branches_data) + 2
                ws1.cell(row=total_row_idx, column=1, value='PROMEDIO / TOTAL')
                ws1.cell(row=total_row_idx, column=2, value=total_employees).number_format = number_format_integer
                ws1.cell(row=total_row_idx, column=3, value=avg_efficiency / 100).number_format = number_format_percent_1dec
                ws1.cell(row=total_row_idx, column=4, value=avg_punctuality / 100).number_format = number_format_percent_1dec
                ws1.cell(row=total_row_idx, column=5, value=avg_sic).number_format = number_format_decimal_2dec
                ws1.cell(row=total_row_idx, column=6, value=total_absences).number_format = number_format_integer
                
                for col_idx in range(1, len(headers1) + 1):
                    cell = ws1.cell(row=total_row_idx, column=col_idx)
                    cell.fill = total_fill; cell.font = total_font; cell.border = thin_border
                        
                column_widths = [25, 12, 15, 15, 15, 12]
                for i, width in enumerate(column_widths, 1): ws1.column_dimensions[get_column_letter(i)].width = width

            if employee_summary_data:
                ws2 = wb.create_sheet(title="Resumen Horas Emp.")
                headers2 = ['ID', 'Empleado', 'Hrs. Trabajadas', 'Hrs. Planificadas', 'Variación', 'Retardos', 'Ausencias']
                ws2.append(headers2)
                
                for col_idx, header in enumerate(headers2, 1):
                    cell = ws2.cell(row=1, column=col_idx)
                    cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align; cell.border = thin_border
                    
                for row_idx, emp in enumerate(employee_summary_data, 2):
                    ws2.cell(row=row_idx, column=1, value=emp.get('ID', 'N/A'))
                    ws2.cell(row=row_idx, column=2, value=emp.get('Empleado', 'Sin Nombre'))
                    ws2.cell(row=row_idx, column=3, value=emp.get('Hrs. Trabajadas', '0:00:00'))
                    ws2.cell(row=row_idx, column=4, value=emp.get('Hrs. Planificadas', '0:00:00'))
                    ws2.cell(row=row_idx, column=5, value=emp.get('Variación', '0:00:00'))
                    
                    cell_retardos = ws2.cell(row=row_idx, column=6, value=emp.get('Retardos', 0))
                    cell_retardos.number_format = number_format_integer
                    
                    cell_ausencias = ws2.cell(row=row_idx, column=7, value=emp.get('Ausencias', 0))
                    cell_ausencias.number_format = number_format_integer
                    
                    for c in range(1, len(headers2) + 1):
                        ws2.cell(row=row_idx, column=c).border = thin_border
                        
                column_widths2 = [10, 40, 18, 18, 15, 12, 12]
                for i, width in enumerate(column_widths2, 1): ws2.column_dimensions[get_column_letter(i)].width = width


            if employee_kpis_data:
                ws3 = wb.create_sheet(title="KPIs Empleado")
                headers3 = ['ID', 'Nombre', 'Tasa Ausentismo (%)', 'Índice Puntualidad (%)', 'Eficiencia Horas (%)', 'SIC']
                ws3.append(headers3)
                
                for col_idx, header in enumerate(headers3, 1):
                    cell = ws3.cell(row=1, column=col_idx)
                    cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align; cell.border = thin_border

                ws3.insert_rows(1); ws3.insert_rows(1)
                title_cell = ws3.cell(row=1, column=1, value="Análisis de KPIs por Empleado")
                subtitle_cell = ws3.cell(row=2, column=1, value="Indicadores de Rendimiento Individual")
                green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True, size=12)
                title_cell.fill = green_fill; title_cell.font = white_font
                subtitle_cell.fill = green_fill; subtitle_cell.font = white_font
                ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers3)); ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers3))


                for row_idx, emp in enumerate(employee_kpis_data, 4):
                    emp_id = emp.get('ID', 'N/A'); emp_nombre = emp.get('Nombre', 'Sin Nombre')
                    tasa_ausentismo_val = emp.get('Tasa Ausentismo (%)', 0) / 100 if emp.get('Tasa Ausentismo (%)') is not None else None
                    puntualidad_val = emp.get('Índice Puntualidad (%)', 0) / 100 if emp.get('Índice Puntualidad (%)') is not None else None
                    eficiencia_val = emp.get('Eficiencia Horas (%)', 0) / 100 if emp.get('Eficiencia Horas (%)') is not None else None
                    sic_val = emp.get('SIC', 0) if emp.get('SIC') is not None else None

                    ws3.cell(row=row_idx, column=1, value=emp_id); ws3.cell(row=row_idx, column=2, value=emp_nombre)
                    c3 = ws3.cell(row=row_idx, column=3, value=tasa_ausentismo_val); c3.number_format = number_format_percent_2dec
                    c4 = ws3.cell(row=row_idx, column=4, value=puntualidad_val); c4.number_format = number_format_percent_2dec
                    c5 = ws3.cell(row=row_idx, column=5, value=eficiencia_val); c5.number_format = number_format_percent_2dec
                    c6 = ws3.cell(row=row_idx, column=6, value=sic_val); c6.number_format = number_format_decimal_2dec

                    for col_idx in range(1, len(headers3) + 1):
                        cell = ws3.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = right_align if col_idx > 2 else left_align
                        if col_idx == 1: cell.alignment = center_align

                        value = cell.value

                        if col_idx == 3: # Tasa Ausentismo
                            if value is not None:
                                if value < 0.05: cell.fill = success_fill; cell.font = success_font
                                elif value <= 0.10: cell.fill = warning_fill; cell.font = warning_font
                                else: cell.fill = danger_fill; cell.font = danger_font
                        elif col_idx == 4: # Puntualidad
                            if value is not None:
                                if value >= 0.95: cell.fill = success_fill; cell.font = success_font
                                elif value >= 0.70: cell.fill = warning_fill; cell.font = warning_font
                                else: cell.fill = danger_fill; cell.font = danger_font
                        elif col_idx == 5: # Eficiencia
                            if value is not None:
                                if value > 1.00: cell.fill = success_fill; cell.font = success_font
                                elif value >= 0.85: cell.fill = warning_fill; cell.font = warning_font
                                else: cell.fill = danger_fill; cell.font = danger_font
                        elif col_idx == 6: # SIC (Usa el valor sin dividir por 100)
                            if sic_val is not None:
                                if sic_val >= 85: cell.fill = success_fill; cell.font = success_font
                                elif sic_val >= 50: cell.fill = warning_fill; cell.font = warning_font
                                else: cell.fill = danger_fill; cell.font = danger_font
                
                legend_start_row = len(employee_kpis_data) + 5
                legend = [
                    ["Interpretación de KPIs:"],
                    ["• Tasa Ausentismo: <5% (Excelente), 5-10% (Aceptable), >10% (Alto)"],
                    ["• Índice Puntualidad: >95 (Excelente), 70-95 (Regular), <70 (Crítico)"],
                    ["• Eficiencia Horas: >100% (Excelente), 85-100% (Regular), <85% (Crítico)"], 
                    ["• SIC: >85 (Excelente), 50-85 (Regular), <50 (Crítico)"] 
                ]
                
                for i, text_list in enumerate(legend):
                    cell = ws3.cell(row=legend_start_row + i, column=1, value=text_list[0])
                    last_col_idx = len(headers3)
                    
                    if i == 0:
                        cell.fill = legend_header_fill; cell.font = legend_header_font; cell.alignment = center_align
                        ws3.merge_cells(start_row=legend_start_row + i, start_column=1, end_row=legend_start_row + i, end_column=last_col_idx)
                    else:
                        cell.font = legend_text_font; cell.alignment = left_align
                        ws3.merge_cells(start_row=legend_start_row + i, start_column=1, end_row=legend_start_row + i, end_column=last_col_idx)

                column_widths3 = [10, 40, 20, 22, 22, 10]
                for i, width in enumerate(column_widths3, 1): ws3.column_dimensions[get_column_letter(i)].width = width

            if not wb.sheetnames:
                return JsonResponse({"error": "No se generaron datos para exportar."}, status=404)
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="dashboard_report.xlsx"'
            return response

        except json.JSONDecodeError as e:
            return JsonResponse({'error': f"Error en formato JSON: {e}"}, status=400)
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': f"Error interno: {e}"}, status=500)
    
    return JsonResponse({'error': 'Método GET no permitido para exportación'}, status=405)

@login_required
def grafica_31pte(request):
    is_admin = request.user.groups.filter(name="Admin").exists()

    context = {
        'is_admin': is_admin, 
    }
    return render(request, "grafica_31pte.html", context)

@login_required
@require_http_methods(["GET"])
def api_dashboard_31pte(request):
    try:
        start_date = request.GET.get("startDate")
        end_date = request.GET.get("endDate")
        if not start_date or not end_date:
            return JsonResponse({"success": False, "error": "Fechas de inicio y fin son requeridas."}, status=400)
        
        resultado = generar_datos_dashboard_31pte(start_date=start_date, end_date=end_date)
        return JsonResponse(resultado)

    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error interno del servidor: {str(e)}"}, status=500)

@login_required
def grafica_villas(request):
    is_admin = request.user.groups.filter(name="Admin").exists()

    context = {
        'is_admin': is_admin, 
    }
    return render(request, "grafica_villas.html", context) 

@login_required
@require_http_methods(["GET"])
def api_dashboard_villas(request):
    try:
        start_date = request.GET.get("startDate")
        end_date = request.GET.get("endDate")
        if not start_date or not end_date:
            return JsonResponse({"success": False, "error": "Fechas de inicio y fin son requeridas."}, status=400)
        
        resultado = generar_datos_dashboard_villas(start_date=start_date, end_date=end_date)
        return JsonResponse(resultado)

    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error interno del servidor: {str(e)}"}, status=500)

@login_required
def grafica_nave(request):
    is_admin = request.user.groups.filter(name="Admin").exists()

    context = {
        'is_admin': is_admin, 
    }
    return render(request, "grafica_nave.html", context) 

@login_required
@require_http_methods(["GET"])
def api_dashboard_nave(request):
    try:
        start_date = request.GET.get("startDate")
        end_date = request.GET.get("endDate")
        if not start_date or not end_date:
            return JsonResponse({"success": False, "error": "Fechas de inicio y fin son requeridas."}, status=400)
        
        resultado = generar_datos_dashboard_nave(start_date=start_date, end_date=end_date)
        return JsonResponse(resultado)

    except Exception as e:
        return JsonResponse({"success": False, "error": f"Error interno del servidor: {str(e)}"}, status=500)

@login_required
@require_http_methods(["GET"])
def api_lista_sucursales(request):
    try:
        sucursales = Sucursal.objects.all()
        data = [
            {"id": s.sucursal_id, "nombre": s.nombre_sucursal} 
            for s in sucursales
        ]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({"error": f"Error en API sucursales: {str(e)}"}, status=500)


@login_required
@require_http_methods(["GET"])
def api_lista_horarios(request):
    try:
        horarios = Horario.objects.all().order_by('hora_entrada')
        
        data = [
            {
                "id": h.horario_id,
                "texto": f"{h.hora_entrada.strftime('%H:%M')} - {h.hora_salida.strftime('%H:%M')}"
                + (f" ({h.descripcion_horario})" if h.descripcion_horario else ""),
                "es_flexible": True if h.descripcion_horario and not h.descripcion_horario.replace(':', '').isdigit() else False
            }
            for h in horarios
        ]
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({"error": f"Error en API horarios: {str(e)}"}, status=500)

@login_required
@require_http_methods(["GET"])
def get_horarios_empleado(request, empleado_id):
    try:
        empleado = get_object_or_404(Empleado.all_objects, empleado_id=empleado_id)
        
        asignaciones = AsignacionHorario.objects.filter(
            empleado=empleado
        ).select_related('sucursal', 'horario', 'dia_especifico', 'tipo_turno')
        
        # Cargamos todos los días (1=Lunes ... 7=Domingo)
        todos_los_dias = list(DiaSemana.objects.all().order_by('dia_id'))
        
        # 🗺️ MAPA DE DECODIFICACIÓN (Igual que tu SQL)
        # X es Miércoles en tu sistema
        mapa_dias = {
            'L': 1, 'M': 2, 'X': 3, 'J': 4, 'V': 5, 'S': 6, 'D': 7
        }

        grupos_de_horarios = {}

        for a in asignaciones:
            if not a.sucursal:
                continue

            # =========================================================
            # 🧩 DECODIFICADOR DE TURNOS (Lógica de tu SQL portada a Python)
            # =========================================================
            dias_permitidos = [] # Aquí guardaremos los IDs de días (1, 2, 3...)

            # CASO A: Tiene un Día Específico (Prioridad Máxima - Caso Rocío)
            if a.dia_especifico:
                dias_permitidos.append(a.dia_especifico.dia_id)
            
            # CASO B: Es un Turno General (L-V, X,J,V, etc.)
            elif a.tipo_turno:
                codigo = a.tipo_turno.descripcion.upper().replace(" ", "") # Ej: "L-V" o "X,J,V"
                
                # Sub-caso 1: Rangos con guión (Ej: "L-V" o "L-J")
                if '-' in codigo:
                    try:
                        inicio_letra, fin_letra = codigo.split('-')
                        inicio = mapa_dias.get(inicio_letra, 1)
                        fin = mapa_dias.get(fin_letra, 7)
                        # Creamos el rango (ej. de 1 a 5)
                        dias_permitidos = list(range(inicio, fin + 1))
                    except:
                        # Si falla, por seguridad damos Lunes a Sábado
                        dias_permitidos = [1, 2, 3, 4, 5, 6]

                # Sub-caso 2: Listas con comas o letras juntas (Ej: "L,X,V" o "LXV")
                else:
                    for letra, id_dia in mapa_dias.items():
                        if letra in codigo:
                            dias_permitidos.append(id_dia)
            
            # CASO C: No tiene nada (Backup de seguridad)
            else:
                dias_permitidos = [1, 2, 3, 4, 5, 6] # L-S por defecto

            # =========================================================
            # FIN DEL DECODIFICADOR
            # =========================================================

            # Definir Texto del Horario
            if a.horario:
                h_id = a.horario.horario_id
                ent_str = a.horario.hora_entrada.strftime('%H:%M')
                sal_str = a.horario.hora_salida.strftime('%H:%M')
                desc = f" ({a.horario.descripcion_horario})" if a.horario.descripcion_horario else ""
            
            elif getattr(a, 'hora_entrada_especifica', None) and getattr(a, 'hora_salida_especifica', None):
                h_id = f"custom_{a.hora_entrada_especifica}_{a.hora_salida_especifica}"
                ent_str = a.hora_entrada_especifica.strftime('%H:%M')
                sal_str = a.hora_salida_especifica.strftime('%H:%M')
                desc = " (Específico)"
            else:
                continue

            texto_horario = f"{ent_str} - {sal_str}{desc}"

            # Agrupar
            key = (a.sucursal.sucursal_id, h_id)
            
            if key not in grupos_de_horarios:
                grupos_de_horarios[key] = {
                    "sucursal_id": a.sucursal.sucursal_id,
                    "sucursal_text": a.sucursal.nombre_sucursal,
                    "horario_id": h_id, 
                    "horario_text": texto_horario,
                    "dias": []
                }
            
            # Asignar Días (Solo los que pasaron el filtro del Decodificador)
            for dia in todos_los_dias:
                if dia.dia_id in dias_permitidos:
                    if dia not in grupos_de_horarios[key]["dias"]:
                        grupos_de_horarios[key]["dias"].append(dia)

        # Generar JSON final
        lista_de_horarios_json = []
        for grupo in grupos_de_horarios.values():
            dias_ordenados = sorted(grupo["dias"], key=lambda d: d.dia_id)
            
            lista_de_horarios_json.append({
                "sucursal_id": grupo["sucursal_id"],
                "sucursal_text": grupo["sucursal_text"],
                "horario_id": grupo["horario_id"],
                "horario_text": grupo["horario_text"],
                "dias_ids": [d.dia_id for d in dias_ordenados],
                "dias_nombres": ", ".join([d.nombre_dia for d in dias_ordenados])
            })

        return JsonResponse(lista_de_horarios_json, safe=False)

    except Exception as e:
        traceback.print_exc() 
        return JsonResponse({"error": f"Error interno: {str(e)}"}, status=500)

@login_required
@require_http_methods(["DELETE"])
def api_eliminar_horario_flexible(request, horario_id):
    try:
        horario = Horario.objects.get(pk=horario_id)
        horario.delete()
        return JsonResponse({"success": "Horario eliminado correctamente y lista limpia."}, status=200)
    
    except Horario.DoesNotExist:
        return JsonResponse({"error": "Horario no encontrado."}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Error al eliminar: {str(e)}"}, status=500)


